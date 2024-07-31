import pandas as pd
import subprocess
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Function to get WHOIS data
def get_whois_data(prefix):
    try:
        # Run the whois command and capture the output as bytes
        result = subprocess.run(['whois', prefix], capture_output=True)
        if result.returncode != 0:
            raise Exception(result.stderr.decode(errors='replace'))
        
        # Decode output manually and handle encoding errors
        output = result.stdout.decode('utf-8', errors='replace')

        # Extract fields using regex
        route_match = re.search(r'route:\s*(\S+)', output, re.IGNORECASE)
        descr_match = re.search(r'descr:\s*(.+)', output, re.IGNORECASE)
        country_match = re.search(r'country:\s*(\S+)', output, re.IGNORECASE)
        last_modified_match = re.search(r'last-modified:\s*(\S+)', output, re.IGNORECASE)
        source_match = re.search(r'source:\s*(\S+)', output, re.IGNORECASE)
        origin_match = re.search(r'origin:\s*AS(\d+)', output, re.IGNORECASE)
        mnt_by_match = re.search(r'mnt-by:\s*(\S+)', output, re.IGNORECASE)

        route = route_match.group(1) if route_match else 'N/A'
        descr = descr_match.group(1).strip() if descr_match else 'N/A'
        country = country_match.group(1) if country_match else 'N/A'
        last_modified = last_modified_match.group(1) if last_modified_match else 'N/A'
        source = source_match.group(1) if source_match else 'N/A'
        origin = origin_match.group(1) if origin_match else 'N/A'
        mnt_by = mnt_by_match.group(1) if mnt_by_match else 'N/A'

        route_object_info = (
            f"route: {route}\n"
            f"descr: {descr}\n"
            f"country: {country}\n"
            f"last-modified: {last_modified}\n"
            f"source: {source}\n"
            f"origin: AS{origin}\n"
            f"mnt-by: {mnt_by}"
        )

        return {
            'Prefix': prefix,
            'ASN': origin,
            'Route-Object Information': route_object_info
        }
    except Exception as e:
        print(f"Error retrieving WHOIS data for prefix {prefix}: {e}")
        return {
            'Prefix': prefix,
            'ASN': 'N/A',
            'Route-Object Information': 'N/A'
        }

# Read prefixes from Excel file
input_file = 'input_prefixes.xlsx'
df = pd.read_excel(input_file)

# Assuming the prefixes are in a column named 'Prefix'
prefix_column = 'Prefix'
results = []

# Iterate over each prefix and get WHOIS data
for prefix in df[prefix_column]:
    whois_data = get_whois_data(prefix)
    results.append(whois_data)

# Create a new DataFrame with the results
output_df = pd.DataFrame(results)

# Save the results to a new Excel file with formatting
output_file = 'whois_data_output.xlsx'

# Create a workbook and select the active worksheet
wb = Workbook()
ws = wb.active

# Add headers with formatting
header_font = Font(bold=True, color="000000")
header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

headers = ['Prefix', 'ASN', 'Route-Object Information']
ws.append(headers)

for col in range(1, len(headers) + 1):
    ws.cell(row=1, column=col).font = header_font
    ws.cell(row=1, column=col).fill = header_fill

# Add data to the worksheet
for index, row in output_df.iterrows():
    ws.append(row.tolist())

# Adjust column width and set text wrapping for the "Route-Object Information" column
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

# Set text wrapping for the "Route-Object Information" column
for cell in ws['C']:
    cell.alignment = Alignment(wrap_text=True)

# Set text alignment for "Prefix" and "ASN" columns
for cell in ws['A']:
    cell.alignment = Alignment(horizontal='center', vertical='center')

for cell in ws['B']:
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Save the workbook
wb.save(output_file)

print(f"WHOIS data saved to {output_file}")
